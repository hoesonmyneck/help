# -*- coding: utf-8 -*-
import openpyxl
from collections import defaultdict

print("=" * 70)
print("ФАЙЛ 1: REGION.xlsx")
print("=" * 70)

wb1 = openpyxl.load_workbook(r'c:\Users\Асан\Desktop\projects\help\REGION.xlsx', read_only=True, data_only=True)
ws1 = wb1.active

rows1 = list(ws1.iter_rows(values_only=True))
headers1 = rows1[0]
print(f"\nЗаголовки: {headers1}")
print(f"Всего строк данных: {len(rows1) - 1}")

# Уникальные пары ID ВЫПЛАТЫ + НАИМЕНОВАНИЕ ВЫПЛАТЫ (индексы 2 и 3)
unique_payments = {}
kato_reg_types = defaultdict(set)   # kato_reg -> set of (id, name)
kato_reg_flag1 = defaultdict(set)  # kato_reg -> set of ids with flag='1'

for row in rows1[1:]:
    pay_id = row[2]
    pay_name = row[3]
    kato_reg = row[1]
    flag = row[4]

    if pay_id is not None:
        unique_payments[pay_id] = pay_name

    if kato_reg is not None and pay_id is not None:
        kato_reg_types[kato_reg].add(pay_id)
        if str(flag) == '1':
            kato_reg_flag1[kato_reg].add(pay_id)

print("\n--- Все уникальные ВЫПЛАТЫ (ID -> НАИМЕНОВАНИЕ) ---")
for pid in sorted(unique_payments.keys()):
    print(f"  ID={pid:>4}  {unique_payments[pid]}")

print(f"\nИтого уникальных выплат: {len(unique_payments)}")

print("\n--- Уникальные KATO_REG ---")
all_katos = sorted(kato_reg_types.keys())
print(f"Всего уникальных KATO_REG: {len(all_katos)}")
print(f"Значения: {all_katos}")

print("\n--- Количество типов выплат (флаг='1') по регионам ---")
for kato in sorted(kato_reg_flag1.keys()):
    cnt = len(kato_reg_flag1[kato])
    ids = sorted(kato_reg_flag1[kato])
    print(f"  KATO_REG={kato:>4}  флаг=1: {cnt} выплат  ids={ids}")

wb1.close()

print("\n")
print("=" * 70)
print("ФАЙЛ 2: cbdiapp_payment_info_for_qlik1.xlsx")
print("=" * 70)

wb2 = openpyxl.load_workbook(
    r'c:\Users\Асан\Desktop\projects\help\backend\data\cbdiapp_payment_info_for_qlik1.xlsx',
    read_only=True, data_only=True
)
ws2 = wb2.active

rows2 = list(ws2.iter_rows(values_only=True))
headers2 = rows2[0]
print(f"\nЗаголовки: {list(enumerate(headers2))}")
print(f"Всего строк данных: {len(rows2) - 1}")

# PAY_TYPE_ID (индекс 8) + PAY_TYPE (индекс 9)
unique_pay_types = {}
# CAT_TYPE_ID (индекс 10) + CAT_TYPE (индекс 11)
unique_cat_types = {}
# Маппинг: PAY_TYPE_ID -> set of CAT_TYPE
pay_to_cat = defaultdict(set)

for row in rows2[1:]:
    pay_type_id = row[8]
    pay_type = row[9]
    cat_type_id = row[10]
    cat_type = row[11]

    if pay_type_id is not None:
        unique_pay_types[pay_type_id] = pay_type

    if cat_type_id is not None:
        unique_cat_types[cat_type_id] = cat_type

    if pay_type_id is not None and cat_type_id is not None:
        pay_to_cat[pay_type_id].add((cat_type_id, cat_type))

print("\n--- Все уникальные PAY_TYPE (PAY_TYPE_ID -> PAY_TYPE) ---")
for pid in sorted(unique_pay_types.keys()):
    cats = pay_to_cat.get(pid, set())
    cats_str = ", ".join(f"[{cid}] {cname}" for cid, cname in sorted(cats))
    print(f"  PAY_TYPE_ID={pid:>4}  {unique_pay_types[pid]}")
    if cats_str:
        print(f"              -> CAT_TYPEs: {cats_str}")

print(f"\nИтого уникальных PAY_TYPE: {len(unique_pay_types)}")

print("\n--- Все уникальные CAT_TYPE (CAT_TYPE_ID -> CAT_TYPE) ---")
for cid in sorted(unique_cat_types.keys()):
    print(f"  CAT_TYPE_ID={cid:>4}  {unique_cat_types[cid]}")

print(f"\nИтого уникальных CAT_TYPE: {len(unique_cat_types)}")

wb2.close()
print("\nГотово.")
