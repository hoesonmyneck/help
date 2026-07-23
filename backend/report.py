"""Выгрузка отчёта: 4 листа.
  1. Аналитика по видам помощи — по принятым заявлениям
  2. Аналитика по видам помощи — по фактической выплате
  3. Аналитика по регионам — по принятым заявлениям
  4. Аналитика по регионам — по фактической выплате
На каждом листе два показателя (кол-во и сумма), у каждого — колонка-итог и разбивка
«В том числе» по категориям благосостояния (ЦКС: A, B, C, …).
Итоговые колонки считаются по всем записям (в т.ч. без ЦКС) — совпадают с сайтом;
разбивка «В том числе» — только по заполненному ЦКС (столбец «Без ЦКС» скрыт).
Бюджет и пол/возраст в отчёт не выгружаются. Форматы: XLSX (openpyxl) и PDF (reportlab).
"""
import io

from sqlalchemy import func, distinct, case as sa_case
from sqlalchemy.orm import Session

from database import Payment
from load_data import (all_region_katos, REGION_NAMES, raion_names_ref,
                       raion_reg_ref, pay_type_names)

NO_CKS = "Без ЦКС"

# Показатели листов: (подпись, индекс в 4-кортеже метрик, формат)
# кортеж метрик = (кол-во заявлений, сумма заявок, факт. получателей, сумма выплат)
APPS_METRICS = [("Принятые заявки", 0, "int"), ("Сумма заявок, ₸", 1, "money")]
PAY_METRICS = [("Фактически выплачено", 2, "int"), ("Сумма выплат, ₸", 3, "money")]


# ─────────────────────────── helpers ───────────────────────────
def _metrics():
    """4 агрегата в фиксированном порядке: заявлений, сумма заявок, факт. получателей, сумма выплат."""
    return [
        func.count(Payment.id),
        func.sum(Payment.dec_pay_sum),
        func.count(distinct(sa_case((Payment.deliv_sum > 0, Payment.sicid)))),
        func.sum(Payment.deliv_sum),
    ]


def _tup(row, off):
    """Достаём метрики из строки, начиная со смещения off."""
    return (int(row[off] or 0), float(row[off + 1] or 0),
            int(row[off + 2] or 0), float(row[off + 3] or 0))


def _norm_cat(v):
    if v is None:
        return NO_CKS
    s = str(v).strip().upper()
    return s or NO_CKS


def _order_cats(cats):
    named = sorted(c for c in cats if c != NO_CKS)
    if NO_CKS in cats:
        named.append(NO_CKS)
    return named


def _tc(s):
    """Регион/район из справочника приходят КАПСОМ — первая заглавная, остальное строчное.
    Города вида «Г.ШЫМКЕНТ» приводим к «г.Шымкент» (строчный маркер города + заглавное название)."""
    s = (s or "").strip()
    if not s:
        return s
    base = s[:1].upper() + s[1:].lower()
    if len(base) > 2 and base[:2] == "Г." and base[2:3].isalpha():
        return "г." + base[2].upper() + base[3:]
    return base


def _all_raions_for_region(region_id, db_katos):
    ref = {d for d, r in raion_reg_ref.items() if r == str(region_id)}
    combined = ref | set(db_katos)
    return sorted(combined, key=lambda k: (int(k) if k.isdigit() else 0, k))


def _only_cks(q):
    """Только записи с заполненным ЦКС — записи без категории в разбивку не идут."""
    return q.filter(Payment.sdu_tzhs.isnot(None), func.trim(Payment.sdu_tzhs) != '')


def _aggregate(db, dim_col, region_filter):
    """Возвращает (totals, percat, cats):
    totals[key]=(4 метрики) сгруппировано по dim; percat[key][cat]=(4 метрики)."""
    # «Итоги» по строке считаем по ВСЕМ записям (в т.ч. без ЦКС) — чтобы совпадало с сайтом
    tq = db.query(dim_col, *_metrics())
    if region_filter is not None:
        tq = tq.filter(Payment.kato_region == region_filter)
    totals = {str(r[0]): _tup(r, 1) for r in tq.group_by(dim_col).all() if r[0] is not None}

    # Разбивка по категориям — только с заполненным ЦКС (столбец «Без ЦКС» не показываем)
    cq = _only_cks(db.query(dim_col, func.upper(Payment.sdu_tzhs), *_metrics()))
    if region_filter is not None:
        cq = cq.filter(Payment.kato_region == region_filter)
    percat, cats = {}, set()
    for r in cq.group_by(dim_col, func.upper(Payment.sdu_tzhs)).all():
        if r[0] is None:
            continue
        cat = _norm_cat(r[1])
        cats.add(cat)
        percat.setdefault(str(r[0]), {})[cat] = _tup(r, 2)
    return totals, percat, cats


def _grand(db, region_filter):
    """Итоговая строка (Республика/регион): (метрики, percat)."""
    gq = db.query(*_metrics())
    if region_filter is not None:
        gq = gq.filter(Payment.kato_region == region_filter)
    total = _tup(gq.one(), 0)

    gcq = _only_cks(db.query(func.upper(Payment.sdu_tzhs), *_metrics()))
    if region_filter is not None:
        gcq = gcq.filter(Payment.kato_region == region_filter)
    gpercat = {_norm_cat(r[0]): _tup(r, 1) for r in gcq.group_by(func.upper(Payment.sdu_tzhs)).all()}
    return total, gpercat


def _row(label, tot, percat_for_key, cats):
    """tot — 4-кортеж «Итоги» строки; per_cat[cat] — 4-кортеж по категории."""
    per = {cat: percat_for_key.get(cat, (0, 0.0, 0, 0.0)) for cat in cats}
    return {"label": label, "tot": tot, "per_cat": per}


def scope_label(region_id):
    if region_id is None:
        return "Республика Казахстан"
    return _tc(REGION_NAMES.get(str(region_id)) or f"Регион {region_id}")


# ─────────────────────────── данные аналитик ───────────────────────────
def _regions_data(db, region_id):
    """Строки для аналитики по регионам (или районам выбранной области)."""
    if region_id is None:
        totals, percat, cats = _aggregate(db, Payment.kato_region, None)
        nm = {str(r[0]): r[1] for r in db.query(Payment.kato_region, Payment.kato_regname).distinct().all()
              if r[0] is not None}
        keys = list(all_region_katos)
        namef = lambda k: _tc(nm.get(k) or REGION_NAMES.get(k, k))
        row_label = "Регион"
    else:
        totals, percat, cats = _aggregate(db, Payment.kato_raion, region_id)
        nm = {str(r[0]): r[1] for r in db.query(Payment.kato_raion, Payment.kato_rainame)
              .filter(Payment.kato_region == region_id).distinct().all() if r[0] is not None}
        keys = _all_raions_for_region(region_id, totals.keys())
        namef = lambda k: _tc(nm.get(k) or raion_names_ref.get(k) or f"Район {k}")
        row_label = "Район"

    gtotal, gpercat = _grand(db, region_id)
    cat_list = _order_cats(cats | set(gpercat.keys()))
    rows = [_row(namef(k), totals.get(k, (0, 0.0, 0, 0.0)), percat.get(k, {}), cat_list) for k in keys]
    rows.sort(key=lambda r: r["tot"][0], reverse=True)   # по кол-ву заявлений
    total = _row(scope_label(region_id), gtotal, gpercat, cat_list)
    return {"row_label": row_label, "cats": cat_list, "total": total, "rows": rows}


def _paytypes_data(db, region_id):
    """Строки для аналитики по видам помощи (полный список видов, с нулями)."""
    totals, percat, cats = _aggregate(db, Payment.pay_type_id, region_id)
    nat_totals, _np, nat_cats = _aggregate(db, Payment.pay_type_id, None)   # полный список видов
    nm = {str(r[0]): r[1] for r in db.query(Payment.pay_type_id, Payment.pay_type).distinct().all()
          if r[0] is not None}
    keys = list(nat_totals.keys())
    namef = lambda k: (nm.get(k) or pay_type_names.get(int(k) if str(k).isdigit() else k) or k)

    gtotal, gpercat = _grand(db, region_id)
    cat_list = _order_cats(nat_cats | cats | set(gpercat.keys()))
    rows = [_row(namef(k), totals.get(k, (0, 0.0, 0, 0.0)), percat.get(k, {}), cat_list) for k in keys]
    rows.sort(key=lambda r: r["tot"][1], reverse=True)   # по сумме заявок
    total = _row(scope_label(region_id), gtotal, gpercat, cat_list)
    return {"row_label": "Вид помощи", "cats": cat_list, "total": total, "rows": rows}


def _collapse_cde(data):
    """Схлопываем категории: A и B отдельно, остальные (C, D, E, …) — в одну колонку суммой."""
    cat_list = data["cats"]
    singles = [c for c in cat_list if c in ("A", "B")]
    rest = [c for c in cat_list if c not in ("A", "B")]
    groups = [(c, [c]) for c in singles]
    if rest:
        groups.append((",".join(rest), rest))

    def remap(row):
        per = {}
        for label, srcs in groups:
            tc = td = tf = tv = 0
            td = tv = 0.0
            for s in srcs:
                a, b, cc, dd = row["per_cat"].get(s, (0, 0.0, 0, 0.0))
                tc += a; td += b; tf += cc; tv += dd
            per[label] = (tc, td, tf, tv)
        row["per_cat"] = per

    remap(data["total"])
    for r in data["rows"]:
        remap(r)
    data["cats"] = [label for label, _ in groups]
    return data


def _mk_sheet(tab, title, data, metrics):
    return {"tab": tab, "title": title, "row_label": data["row_label"],
            "cats": data["cats"], "total": data["total"], "rows": data["rows"],
            "metrics": metrics}


def build_sheets(db: Session, region_id):
    """4 листа: виды помощи (заявления, выплаты), регионы (заявления, выплаты)."""
    pt = _collapse_cde(_paytypes_data(db, region_id))
    rg = _collapse_cde(_regions_data(db, region_id))
    return [
        _mk_sheet("Виды помощи (заявления)",
                  "Аналитика по видам помощи (по принятым заявлениям)", pt, APPS_METRICS),
        _mk_sheet("Виды помощи (выплаты)",
                  "Аналитика по видам помощи (по фактической выплате)", pt, PAY_METRICS),
        _mk_sheet("Регионы (заявления)",
                  "Аналитика по регионам (по принятым заявлениям)", rg, APPS_METRICS),
        _mk_sheet("Регионы (выплаты)",
                  "Аналитика по регионам (по фактической выплате)", rg, PAY_METRICS),
    ]


# ─────────────────────────── XLSX ───────────────────────────
_CAT_FILLS = ["C8E6C9", "A5D6A7", "FFE0B2", "FFCCBC", "B3E5FC", "D1C4E9", "F0F4C3", "F8BBD0"]
_FMT = {"int": "#,##0", "money": "#,##0"}


def write_xlsx(sheets, scope):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)

    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")
    hdr_fill = PatternFill("solid", fgColor="ECEFF1")
    metric_fill = PatternFill("solid", fgColor="CFD8DC")
    totalrow_fill = PatternFill("solid", fgColor="FFF9C4")
    hdr_font = Font(bold=True, size=10)
    bold = Font(bold=True)

    for sh in sheets:
        cats = sh["cats"]
        metrics = sh["metrics"]
        ncat = len(cats)
        block = 1 + ncat                       # 1 итог-колонка + «В том числе» (ncat)
        ncols = 1 + len(metrics) * block
        ws = wb.create_sheet(title=sh["tab"][:31])

        ws.cell(1, 1, f'{sh["title"]} — {scope}').font = Font(bold=True, size=12)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)

        r1, r2 = 2, 3
        c = ws.cell(r1, 1, sh["row_label"])
        c.font = hdr_font; c.alignment = center; c.fill = hdr_fill; c.border = border
        ws.merge_cells(start_row=r1, start_column=1, end_row=r2, end_column=1)

        for mi, (mlabel, midx, mfmt) in enumerate(metrics):
            start = 2 + mi * block
            # колонка-итог показателя (вертикально объединяем обе строки шапки)
            tc = ws.cell(r1, start, mlabel)
            tc.font = hdr_font; tc.alignment = center; tc.fill = metric_fill; tc.border = border
            ws.merge_cells(start_row=r1, start_column=start, end_row=r2, end_column=start)
            # «В том числе» над категориями
            gc = ws.cell(r1, start + 1, "В том числе")
            gc.font = hdr_font; gc.alignment = center; gc.fill = hdr_fill; gc.border = border
            ws.merge_cells(start_row=r1, start_column=start + 1, end_row=r1, end_column=start + ncat)
            for ci, cat in enumerate(cats):
                cell = ws.cell(r2, start + 1 + ci, cat)
                cell.font = hdr_font; cell.alignment = center; cell.border = border
                cell.fill = PatternFill("solid", fgColor=_CAT_FILLS[ci % len(_CAT_FILLS)])

        def write_row(rownum, rd, is_total):
            lab = ws.cell(rownum, 1, rd["label"])
            lab.alignment = left; lab.border = border
            if is_total:
                lab.font = bold
            for mi, (mlabel, midx, mfmt) in enumerate(metrics):
                start = 2 + mi * block
                fmt = _FMT[mfmt]
                tcell = ws.cell(rownum, start, rd["tot"][midx])
                tcell.number_format = fmt; tcell.alignment = right; tcell.border = border
                if is_total:
                    tcell.font = bold
                for ci, cat in enumerate(cats):
                    cell = ws.cell(rownum, start + 1 + ci, rd["per_cat"][cat][midx])
                    cell.number_format = fmt; cell.alignment = right; cell.border = border
                    if is_total:
                        cell.font = bold
            if is_total:
                for col in range(1, ncols + 1):
                    ws.cell(rownum, col).fill = totalrow_fill

        cur = r2 + 1
        write_row(cur, sh["total"], True); cur += 1
        for rd in sh["rows"]:
            write_row(cur, rd, False); cur += 1

        ws.column_dimensions["A"].width = 36
        for col in range(2, ncols + 1):
            fmt = metrics[(col - 2) // block][2]
            ws.column_dimensions[get_column_letter(col)].width = 15 if fmt == "money" else 10
        ws.freeze_panes = "B4"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────── PDF ───────────────────────────
def _fmt_int(v):
    return f"{int(v):,}".replace(",", " ")


def _fmt_money(v):
    return f"{int(round(v)):,}".replace(",", " ")


def _fmt_val(v, fmt):
    return _fmt_int(v) if fmt == "int" else _fmt_money(v)


def _register_font():
    """Регистрируем шрифт с кириллицей. Возвращает (regular, bold) имена шрифтов."""
    import os
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    candidates = [
        ("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
         "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
        ("C:/Windows/Fonts/arial.ttf", "C:/Windows/Fonts/arialbd.ttf"),
    ]
    for reg, bold in candidates:
        if os.path.exists(reg):
            try:
                if "RPT" not in pdfmetrics.getRegisteredFontNames():
                    pdfmetrics.registerFont(TTFont("RPT", reg))
                    pdfmetrics.registerFont(TTFont("RPT-Bold", bold if os.path.exists(bold) else reg))
                return "RPT", "RPT-Bold"
            except Exception:
                pass
    return "Helvetica", "Helvetica-Bold"


def write_pdf(sheets, scope):
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, PageBreak)

    FONT, FONT_B = _register_font()
    styles = getSampleStyleSheet()
    h_style = ParagraphStyle("h", parent=styles["Title"], fontName=FONT_B, fontSize=12, spaceAfter=6)
    cell_style = ParagraphStyle("c", fontName=FONT, fontSize=7, leading=8.4)

    cat_hexes = ["#C8E6C9", "#A5D6A7", "#FFE0B2", "#FFCCBC", "#B3E5FC",
                 "#D1C4E9", "#F0F4C3", "#F8BBD0"]
    label_w = 130
    int_head_w = 90   # столбец-итог количества («Принятые заявки»…) — под заголовок в одну строку
    int_cat_w = 40    # разбивка количества (A/B/C,D,E) — узкая
    money_w = 64      # денежные столбцы (₸) — уже, чем были

    blocks = []
    for si, sh in enumerate(sheets):
        cats = sh["cats"]
        metrics = sh["metrics"]
        ncat = len(cats)
        block = 1 + ncat

        def P(txt):
            return Paragraph(str(txt), cell_style)

        # шапка (2 строки)
        head1 = [sh["row_label"]]
        head2 = [""]
        for (mlabel, midx, mfmt) in metrics:
            head1 += [mlabel, "В том числе"] + [""] * (ncat - 1)
            head2 += [""] + list(cats)
        data = [[P(x) for x in head1], [P(x) for x in head2]]

        def make_row(rd):
            r = [P(rd["label"])]
            for (mlabel, midx, mfmt) in metrics:
                r.append(_fmt_val(rd["tot"][midx], mfmt))
                for cat in cats:
                    r.append(_fmt_val(rd["per_cat"][cat][midx], mfmt))
            return r

        data.append(make_row(sh["total"]))
        for rd in sh["rows"]:
            data.append(make_row(rd))

        col_widths = [label_w]
        for (mlabel, midx, mfmt) in metrics:
            if mfmt == "money":
                col_widths += [money_w] * (1 + ncat)
            else:
                col_widths += [int_head_w] + [int_cat_w] * ncat
        t = Table(data, colWidths=col_widths, repeatRows=2)

        ts = [
            ("FONTNAME", (0, 0), (-1, -1), FONT),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("FONTNAME", (0, 0), (-1, 1), FONT_B),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("ALIGN", (0, 0), (0, -1), "LEFT"),
            ("ALIGN", (1, 0), (-1, 1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
            ("SPAN", (0, 0), (0, 1)),
            ("BACKGROUND", (0, 0), (-1, 1), colors.HexColor("#ECEFF1")),
            ("BACKGROUND", (0, 2), (-1, 2), colors.HexColor("#FFF9C4")),
            ("FONTNAME", (0, 2), (-1, 2), FONT_B),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ]
        for mi, (mlabel, midx, mfmt) in enumerate(metrics):
            tcol = 1 + mi * block
            ts.append(("SPAN", (tcol, 0), (tcol, 1)))                    # колонка-итог (вертикально)
            ts.append(("SPAN", (tcol + 1, 0), (tcol + ncat, 0)))         # «В том числе»
            ts.append(("BACKGROUND", (tcol, 0), (tcol, 1), colors.HexColor("#CFD8DC")))
            for ci in range(ncat):
                ts.append(("BACKGROUND", (tcol + 1 + ci, 1), (tcol + 1 + ci, 1),
                           colors.HexColor(cat_hexes[ci % len(cat_hexes)])))
        t.setStyle(TableStyle(ts))

        blocks.append((Paragraph(f'{sh["title"]}<br/>{scope}', h_style), t))

    # ширина страницы — по самому широкому листу
    def sheet_w(sh):
        w = label_w + 30
        ncat = len(sh["cats"])
        for (mlabel, midx, mfmt) in sh["metrics"]:
            if mfmt == "money":
                w += money_w * (1 + ncat)
            else:
                w += int_head_w + int_cat_w * ncat
        return w
    max_w = max(sheet_w(sh) for sh in sheets)
    page_w = max_w
    l_m = r_m = 14
    t_m = b_m = 16
    avail_w = page_w - l_m - r_m

    # высоту листа берём по реальной высоте самой большой таблицы (reportlab.wrap),
    # чтобы каждая таблица помещалась на одну страницу без переноса
    def block_h(title, tbl):
        th = title.wrap(avail_w, 100000)[1]
        bh = tbl.wrap(avail_w, 100000)[1]
        return th + bh + 8
    page_h = max(block_h(title, tbl) for title, tbl in blocks) + t_m + b_m + 12

    story = []
    for i, (title, tbl) in enumerate(blocks):
        if i > 0:
            story.append(PageBreak())
        story.append(title)
        story.append(tbl)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=(page_w, page_h),
                            leftMargin=l_m, rightMargin=r_m, topMargin=t_m, bottomMargin=b_m)
    doc.build(story)
    return buf.getvalue()
