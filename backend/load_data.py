import os
import re
from datetime import datetime, date
from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.orm import Session
from database import engine, Base, Payment, BudgetItem, RegionBudget

region_help_ids: dict[str, set] = {}   # {kato_reg_str: {pay_type_id, ...}} where flag='1'
raion_help_ids:  dict[str, set] = {}   # {kato_dis_str: {pay_type_id, ...}} where flag='1'
all_region_katos: list[str] = []       # all KATO_REG values from REGION.xlsx in order
pay_type_names:  dict[int, str] = {}   # {pay_type_id: НАИМЕНОВАНИЕ ВЫПЛАТЫ} from REGION.xlsx

# Region/raion display names by KATO code — populated from the reference files
REGION_NAMES:    dict[str, str] = {}   # {kato_reg: REG name} from REGION.xlsx
raion_names_ref: dict[str, str] = {}   # {kato_dis: DIS name} from RAION.xlsx
raion_reg_ref:   dict[str, str] = {}   # {kato_dis: kato_reg} — region that owns each raion

# Payment settings rows from PAYMENT_SETTING_FOR_QLIK.xlsx
# each item: (kato_reg_str, kato_dis_str, pay_type_id, pay_name, cat_type, max_val|None)
settings_rows: list = []
settings_pay_names: dict[int, str] = {}  # {pay_type_id: НАИМЕНОВАНИЕ ВЫПЛАТЫ} from settings file

# ── Раздел «Всеобуч»: каталог видов услуг, выведенный из vseobuch.xlsx ──
# У всеобуча нет отдельного settings-файла, поэтому «каталог» строится из самих данных.
# Все структуры мутируются на месте (не переприсваиваются) — импорт по ссылке в main.py.
vseobuch_settings_rows: list = []            # аналог settings_rows для всеобуча
vseobuch_settings_pay_names: dict[int, str] = {}   # {pid: SERVICE_NAME}
vseobuch_pay_type_names: dict[int, str] = {}       # аналог pay_type_names (REGION.xlsx)

# Regions that exist on the map / administratively but provide nothing (not in REGION.xlsx)
EXTRA_REGIONS = [('10', 'АБАЙСКАЯ ОБЛАСТЬ'), ('62', 'УЛЫТАУСКАЯ ОБЛАСТЬ')]


def _to_kato_str(val) -> str | None:
    if val is None:
        return None
    try:
        return str(int(float(val)))
    except (ValueError, TypeError):
        return str(val).strip() or None


def load_reference_data():
    base = os.path.join(os.path.dirname(__file__), "data")

    wb = load_workbook(os.path.join(base, "REGION.xlsx"), read_only=True, data_only=True)
    ws = wb.active
    region_help_ids.clear()
    pay_type_names.clear()
    all_region_katos.clear()
    REGION_NAMES.clear()
    katos_seen: list[str] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        # cols: 0=KATO_REG, 1=PAY_TYPE_ID, 2=PAY_TYPE(name), 3=FLAG_ID, 4=FLAG(text), 5=REG(name)
        kato = _to_kato_str(row[0])
        pay_id = int(row[1]) if row[1] is not None else None
        name = str(row[2]).strip() if row[2] is not None else None
        flag = str(row[3]).strip() if row[3] is not None else ''
        reg_name = str(row[5]).strip() if row[5] is not None else None
        if kato and kato not in katos_seen:
            katos_seen.append(kato)
        if kato and reg_name:
            REGION_NAMES.setdefault(kato, reg_name)
        if pay_id and flag == '1' and kato:
            region_help_ids.setdefault(kato, set()).add(pay_id)
        if pay_id and name and pay_id not in pay_type_names:
            pay_type_names[pay_id] = name
    all_region_katos.extend(katos_seen)
    # Add administrative regions absent from REGION.xlsx (Abai, Ulytau) — they provide nothing
    for extra_kato, extra_name in EXTRA_REGIONS:
        if extra_kato not in all_region_katos:
            all_region_katos.append(extra_kato)
        REGION_NAMES.setdefault(extra_kato, extra_name)
    wb.close()

    wb = load_workbook(os.path.join(base, "RAION.xlsx"), read_only=True, data_only=True)
    ws = wb.active
    raion_help_ids.clear()
    raion_names_ref.clear()
    raion_reg_ref.clear()
    for row in ws.iter_rows(min_row=2, values_only=True):
        # cols: 0=KATO_REG, 1=KATO_DIS, 2=PAY_TYPE_ID, 3=PAY_TYPE, 4=FLAG_ID, 5=FLAG, 6=REG, 7=DIS(name)
        reg_kato = _to_kato_str(row[0])
        kato = _to_kato_str(row[1])
        pay_id = int(row[2]) if row[2] is not None else None
        flag = str(row[4]).strip() if row[4] is not None else ''
        dis_name = str(row[7]).strip() if row[7] is not None else None
        if kato and dis_name:
            raion_names_ref.setdefault(kato, dis_name)
        if kato and reg_kato:
            raion_reg_ref.setdefault(kato, reg_kato)
        if kato and pay_id and flag == '1':
            raion_help_ids.setdefault(kato, set()).add(pay_id)
    wb.close()

    # Payment settings (виды помощи / категории людей with configured max sum)
    settings_rows.clear()
    settings_pay_names.clear()
    settings_path = os.path.join(base, "PAYMENT_SETTING_FOR_QLIK.xlsx")
    if os.path.exists(settings_path):
        wb = load_workbook(settings_path, read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            # cols: 0=KATO_REG, 1=KATO_DIS, 2=ID ВЫПЛАТЫ, 3=НАИМЕНОВАНИЕ,
            # 4=ID КАТЕГОРИЙ, 5=КАТЕГОРИИ ПОЛУЧАТЕЛЕЙ, 6=ПЕРИОДИЧНОСТЬ,
            # 7=МАКСИМАЛЬНАЯ СУММА ВЫПЛАТЫ, 8=UNIT_ID
            kato_reg = _to_kato_str(row[0])
            kato_dis = _to_kato_str(row[1])
            pay_id = int(row[2]) if row[2] is not None else None
            pay_name = str(row[3]).strip() if row[3] is not None else None
            cat = str(row[5]).strip() if row[5] is not None else None
            max_sum = row[7]
            max_val = None
            try:
                if max_sum is not None and str(max_sum).strip() not in ('', '-'):
                    max_val = float(max_sum)
            except (ValueError, TypeError):
                max_val = None
            if pay_id and pay_name and pay_id not in settings_pay_names:
                settings_pay_names[pay_id] = pay_name
            if pay_id and cat:
                # (kato_reg, kato_dis, pay_id, pay_name, cat, max_val|None)
                settings_rows.append((kato_reg, kato_dis, pay_id, pay_name, cat, max_val))
        wb.close()

    print(f"Reference loaded: {len(region_help_ids)} regions, "
          f"{len(raion_help_ids)} raions, {len(settings_rows)} settings rows")


def parse_date(val):
    if val is None:
        return None
    if isinstance(val, (date, datetime)):
        return val if isinstance(val, date) else val.date()
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def parse_datetime(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%d.%m.%Y %H:%M:%S", "%Y-%m-%d %H:%M:%S", "%d.%m.%Y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    return None


def parse_num(val, cast=float):
    if val is None:
        return None
    try:
        return cast(val)
    except (ValueError, TypeError):
        return None


def _fmt_geo_name(raw):
    """Привести REG/DIS (ВЕРХНИЙ РЕГИСТР) к аккуратному виду для отображения.

    «КАРАГАНДА Г.А.»          -> «г.Караганда»   (город-администрация)
    «Г.АСТАНА» / «Г.КОСШЫ»    -> «г.Астана» / «г.Косшы»
    «САРЫСУСКИЙ РАЙОН»        -> «Сарысуский район»
    «КАРАГАНДИНСКАЯ ОБЛАСТЬ»  -> «Карагандинская область»
    """
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    # заглавная первая буква и буквы после дефиса: «западно-казахстанская» -> «Западно-Казахстанская»
    def _cap(x):
        return re.sub(r'(^|-)([а-яёa-z])', lambda mm: mm.group(1) + mm.group(2).upper(), x.lower())
    m = re.match(r'^(.*?)\s+г\.а\.?$', s, re.IGNORECASE)      # «X Г.А.»
    if m:
        return 'г.' + _cap(m.group(1).strip())
    m = re.match(r'^г\.\s*(.+)$', s, re.IGNORECASE)           # «Г.X»
    if m:
        return 'г.' + _cap(m.group(1).strip())
    return _cap(s)


def parse_gender(val):
    """New data has gender as text МУЖСКОЙ/ЖЕНСКИЙ; map to 1/2."""
    if val is None:
        return None
    s = str(val).strip().upper()
    if s.startswith('М'):
        return 1
    if s.startswith('Ж'):
        return 2
    return parse_num(val, int)


def _detect_app_id_offset(ws) -> int:
    """Найти индекс колонки APP_ID в строке заголовков.
    Файлы бывают с ведущим столбцом-счётчиком (offset=1) и без него (offset=0)."""
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    for idx, val in enumerate(header):
        if val is not None and str(val).strip().upper() == 'APP_ID':
            return idx
    # запасной вариант: 31 колонка => есть ведущий счётчик
    return 1 if len(header) >= 31 else 0


def parse_payment_rows(ws, dataset: str = "mio") -> list:
    """Распарсить лист в список Payment. Колонки идут фиксированным порядком от APP_ID.
    `dataset` помечает раздел ('mio' или 'vseobuch') — по нему фильтруются все запросы."""
    b = _detect_app_id_offset(ws)
    rows_data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[b] is None:
            continue
        # Регион/район определяем по KATO_REG/REG (кол. 6-7) и KATO_DIS/DIS (кол. 8-9),
        # а НЕ по KATO_REGION/KATO_RAION (кол. 27-28): последние дробят города
        # республиканского значения (Астана, Шымкент) на районы и иногда ошибочно
        # приписывают чужие записи. KATO_REG (9 знаков) // 10_000_000 = 2-значный код
        # региона; KATO_DIS // 100_000 = 4-значный код района.
        _kreg = parse_num(row[b + 5], int)   # KATO_REG
        _kdis = parse_num(row[b + 7], int)   # KATO_DIS
        rows_data.append(Payment(
            dataset=dataset,
            app_id=parse_num(row[b + 0], int),
            app_date=parse_date(row[b + 1]),
            app_date_close=parse_date(row[b + 2]),
            app_status=str(row[b + 3]).strip() if row[b + 3] else None,
            iin=str(row[b + 4]).strip() if row[b + 4] else None,
            kato_reg=_kreg,
            kato_dis=_kdis,
            pay_type_id=parse_num(row[b + 9], int),
            pay_type=str(row[b + 10]).strip() if row[b + 10] else None,
            cat_type_id=parse_num(row[b + 11], int),
            cat_type=str(row[b + 12]).strip() if row[b + 12] else None,
            period=str(row[b + 13]).strip() if row[b + 13] else None,
            unit_id=parse_num(row[b + 14], int),
            max_pay_sum=parse_num(row[b + 15]),
            decision=str(row[b + 16]).strip() if row[b + 16] else None,
            dec_pay_sum=parse_num(row[b + 17]),
            deliv_date=parse_date(row[b + 18]),
            deliv_sum=parse_num(row[b + 19]),
            mrp=parse_num(row[b + 20]),
            sys_date=parse_datetime(row[b + 21]),
            sicid=parse_num(row[b + 22], int),
            gender_id=parse_num(row[b + 23], int),
            vozrast=parse_num(row[b + 24], int),
            sdu_tzhs=str(row[b + 25]).strip() if row[b + 25] else None,
            kato_region=(_kreg // 10_000_000) if _kreg is not None else None,
            kato_raion=(_kdis // 100_000) if _kdis is not None else None,
            kato_regname=_fmt_geo_name(row[b + 6]),
            kato_rainame=_fmt_geo_name(row[b + 8]),
        ))
    return rows_data


def ensure_dataset_column():
    """Идемпотентная миграция: добавить колонку payments.dataset и проставить 'mio'
    существующим строкам (раздел МИО — данные, которые уже загружены)."""
    Base.metadata.create_all(bind=engine)
    with engine.begin() as conn:
        conn.execute(text(
            "ALTER TABLE payments ADD COLUMN IF NOT EXISTS dataset VARCHAR(20)"))
        conn.execute(text(
            "UPDATE payments SET dataset='mio' WHERE dataset IS NULL"))


def replace_payments_from_file(file_obj, dataset: str = "mio") -> int:
    """Загрузить выплаты из переданного xlsx (файл/поток): заменить строки раздела
    `dataset` и залить заново. Возвращает число загруженных строк. Другие разделы
    (напр. всеобуч) не затрагиваются. Используется админ-загрузкой."""
    ensure_dataset_column()
    wb = load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb.active
    rows_data = parse_payment_rows(ws, dataset=dataset)
    wb.close()
    if not rows_data:
        raise ValueError("В файле не найдено ни одной строки данных (нет колонки APP_ID?)")
    with Session(engine) as session:
        session.execute(text("DELETE FROM payments WHERE dataset=:ds"), {"ds": dataset})
        session.add_all(rows_data)
        session.commit()
    print(f"Admin upload: replaced '{dataset}' payments with {len(rows_data)} rows")
    return len(rows_data)


def _vs_nn(v):
    """Значение или None для пустых/`[NULL]`."""
    if v is None:
        return None
    s = str(v).strip()
    if s == '' or s == '[NULL]':
        return None
    return v


def _vs_money(v):
    """Сумма из vseobuch: число или текст вида '50 831' (неразрывный пробел)."""
    v = _vs_nn(v)
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace('\xa0', '').replace(' ', '').replace(',', '.')
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


# Раскладка колонок листа ' social_services' в vseobuch.xlsx (0-индексно):
# 0 счётчик, 1 REG_RU, 2 CODE_KATO, 3 RAI_RU, 4 IDRAI, 5 REGION_NAME, 6 BIN, 7 DEP_NAME,
# 8 IIN, 9 SERVICE_NAME, 10 APP_DATE, 11 SOURCE_NAME, 12 APP_STATUS_NAME, 13 NAZN_SUM,
# 14 PERIOD_TYPE, 15 ED_IZM, 16 STAT_RESHENIYA, 17 PODPISANO, 18 SDU_TZHS,
# 19 SK_FAMILY_ID, 20 GENDER_ID, 21 VOZRAST
def parse_vseobuch_rows(ws):
    """Распарсить лист всеобуча в (список Payment, каталог видов услуг).

    У всеобуча нет числового ID вида услуги — синтезируем стабильный ID из
    SERVICE_NAME. Сумм всего одна (NAZN_SUM = назначенная) — кладём её в
    max/dec/deliv, чтобы метрики «принято/факт» работали как в МИО.
    Категорий получателей нет — считаем каждую услугу отдельной категорией."""
    service_ids: dict[str, int] = {}

    def _svc_id(name):
        if name not in service_ids:
            service_ids[name] = len(service_ids) + 1
        return service_ids[name]

    rows_data = []
    settings_seen: dict = {}   # (reg2, rai4, pid) -> (name, max_val)
    for row in ws.iter_rows(min_row=2, values_only=True):
        counter = parse_num(row[0], int)
        code_kato = parse_num(row[2], int)
        service = _vs_nn(row[9])
        if code_kato is None or service is None:
            continue
        pid = _svc_id(str(service).strip())
        summ = _vs_money(row[13])
        kato_region = code_kato // 10_000_000
        kato_raion = code_kato // 100_000
        rows_data.append(Payment(
            dataset="vseobuch",
            app_id=counter,
            app_date=parse_date(row[10]),
            app_status=str(_vs_nn(row[12])).strip() if _vs_nn(row[12]) else None,
            iin=str(_vs_nn(row[8])).strip() if _vs_nn(row[8]) else None,
            kato_reg=code_kato,
            kato_dis=code_kato,
            pay_type_id=pid,
            pay_type=str(service).strip(),
            cat_type_id=pid,
            cat_type=str(service).strip(),
            period=str(_vs_nn(row[14])).strip() if _vs_nn(row[14]) else None,
            unit_id=None,
            max_pay_sum=summ,
            decision=str(_vs_nn(row[16])).strip() if _vs_nn(row[16]) else None,
            dec_pay_sum=summ,
            deliv_date=None,
            deliv_sum=summ,
            mrp=None,
            sys_date=None,
            sicid=parse_num(row[19], int),
            gender_id=parse_gender(row[20]),
            vozrast=parse_num(row[21], int),
            sdu_tzhs=str(_vs_nn(row[18])).strip() if _vs_nn(row[18]) else None,
            kato_region=kato_region,
            kato_raion=kato_raion,
            kato_regname=_fmt_geo_name(row[1]),
            kato_rainame=_fmt_geo_name(row[3]),
        ))
        key = (str(kato_region), str(kato_raion), pid)
        prev = settings_seen.get(key)
        mx = summ if summ is not None else (prev[1] if prev else 1.0)
        if prev is None or (summ is not None and (prev[1] is None or summ > prev[1])):
            settings_seen[key] = (str(service).strip(), mx)

    # Собрать каталог (мутируем глобальные структуры на месте)
    vseobuch_settings_pay_names.clear()
    vseobuch_pay_type_names.clear()
    for name, pid in service_ids.items():
        vseobuch_settings_pay_names[pid] = name
        vseobuch_pay_type_names[pid] = name
    vseobuch_settings_rows.clear()
    for (reg2, rai4, pid), (name, mx) in settings_seen.items():
        # (kato_reg2, kato_dis4, pid, pay_name, cat=name, max_val) — как settings_rows
        vseobuch_settings_rows.append((reg2, rai4, pid, name, name, mx))

    return rows_data


def load_vseobuch():
    """Построить каталог всеобуча из vseobuch.xlsx и (идемпотентно) залить строки
    в раздел 'vseobuch'. Каталог строится ВСЕГДА (он в памяти), строки — только если
    в БД их ещё нет."""
    ensure_dataset_column()

    path = os.path.join(os.path.dirname(__file__), "data", "vseobuch.xlsx")
    if not os.path.exists(path):
        print("load_vseobuch: vseobuch.xlsx не найден — пропуск")
        return

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_data = parse_vseobuch_rows(ws)   # всегда пересобирает каталог в памяти
    wb.close()

    print(f"Vseobuch catalog: {len(vseobuch_settings_pay_names)} services, "
          f"{len(vseobuch_settings_rows)} settings rows")

    with engine.connect() as conn:
        count = conn.execute(text(
            "SELECT COUNT(*) FROM payments WHERE dataset='vseobuch'")).scalar()
    if count and count > 0:
        return

    if not rows_data:
        print("load_vseobuch: в vseobuch.xlsx не найдено строк данных")
        return

    with Session(engine) as session:
        session.add_all(rows_data)
        session.commit()

    print(f"Loaded {len(rows_data)} rows from vseobuch.xlsx")


def load_excel():
    Base.metadata.create_all(bind=engine)
    ensure_dataset_column()

    with engine.connect() as conn:
        count = conn.execute(text(
            "SELECT COUNT(*) FROM payments WHERE dataset='mio'")).scalar()
        if count and count > 0:
            return

    path = os.path.join(os.path.dirname(__file__), "data", "new14.xlsx")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_data = parse_payment_rows(ws, dataset="mio")
    wb.close()

    with Session(engine) as session:
        session.add_all(rows_data)
        session.commit()

    print(f"Loaded {len(rows_data)} rows from Excel")


def load_budget():
    """Load help_types.xlsx into budget_items table (idempotent).

    KATO_REG and KATO_DIS in help_types.xlsx use the same short codes as
    Payment.kato_region / Payment.kato_raion, so store them as-is (no division).
    Auto-migrates old rows where the division bug set kato_region=0.
    """
    Base.metadata.create_all(bind=engine)

    path = os.path.join(os.path.dirname(__file__), "data", "help_types.xlsx")
    if not os.path.exists(path):
        return

    with engine.connect() as conn:
        count = conn.execute(text("SELECT COUNT(*) FROM budget_items")).scalar()
        if count > 0:
            max_reg = conn.execute(text(
                "SELECT MAX(kato_region) FROM budget_items WHERE kato_region IS NOT NULL"
            )).scalar() or 0
            if max_reg != 0:
                return  # data already correct
            # kato_region is all-zero (old division bug) — delete and reload
            conn.execute(text("DELETE FROM budget_items"))
            conn.commit()

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    rows_data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        # cols: 0=OBLNAME, 1=ID, 2=RNAME, 3=KATO_REG, 4=KATO_DIS, 5=PLANSUM
        kato_reg_raw = parse_num(row[3], int)
        kato_dis_raw = parse_num(row[4], int)
        plansum = parse_num(row[5])
        if kato_reg_raw is None or plansum is None:
            continue
        rows_data.append(BudgetItem(
            pay_type_id=parse_num(row[1], int),
            kato_region=kato_reg_raw,
            kato_raion=kato_dis_raw,
            plansum=plansum,
        ))

    wb.close()

    with Session(engine) as session:
        session.add_all(rows_data)
        session.commit()

    print(f"Loaded {len(rows_data)} budget rows from help_types.xlsx")


def load_region_budget():
    """Load budget.xlsx into region_budgets table (idempotent).

    budget.xlsx structure: col[0]=row#, col[1]=kato_region (2-digit int),
    col[2]=region name, col[3]=budget float.
    """
    Base.metadata.create_all(bind=engine)

    path = os.path.join(os.path.dirname(__file__), "data", "budget.xlsx")
    if not os.path.exists(path):
        return

    with engine.connect() as conn:
        count = conn.execute(text("SELECT COUNT(*) FROM region_budgets")).scalar()
        if count > 0:
            return

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    rows_data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        kato = parse_num(row[1], int)
        name = str(row[2]).strip() if row[2] is not None else None
        budget = parse_num(row[3])
        if kato is None or budget is None:
            continue
        rows_data.append(RegionBudget(kato_region=kato, region_name=name, budget=budget))

    wb.close()

    with Session(engine) as session:
        session.add_all(rows_data)
        session.commit()

    print(f"Loaded {len(rows_data)} region budget rows from budget.xlsx")
